import csv
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import PatternFill
from color_analyse import Color_Class
from color_analyse import Color_Analysis

class Write_Results_Class:
	def Write_To_CSV(fieldnames, data ):
		with open('results.csv', 'w', newline='') as csvfile:
			writer = csv.DictWriter(csvfile, fieldnames = fieldnames)
			writer.writeheader()
			for datarow in data:
				writer.writerow(datarow)

	def Initialize_JS_File(file_name):
		with open(file_name, 'w') as file:
			file.write("'use strict';\n")
			file.close()

	def Initialize_JSON_File(file_name):
		with open(file_name, 'w') as file:
			start = "var allEyeshadowDetails = ["
			file.write(start)
			file.close()

	def Close_JSON_File(file_name):
		with open(file_name, 'a') as file:
			end = "];"
			file.write(end)
			file.close()

	def Write_To_JS(brand, data, file_name):
		with open(file_name, 'a') as file:
			start = "window.{0}".format(brand["name"].replace(" ", "_").replace("'","").replace("+", ""))
			file.write(start)
			file.write("={")

			#write eyeshadow names here
			#using avgrgb might need to do another type later
			for eyeshadow in data:
				eyeshadowline = "'{0}': 'rgb{1}',".format(eyeshadow["Name"].replace("'",""), eyeshadow["AvgRGB"])
				file.write(eyeshadowline)
			file.write("};\n")

			file.close()

	def Write_To_JSon_Objects( data, file_name):
		with open(file_name, 'a') as file:
			# start = "var allEyeshadowDetails = {"
			# file.write(start)
			for eyeshadow in data:
				eyeshadowline = '{{"name":"{0}", "brand":"{1}", "palette": "{2}", "finish": "{3}" }},\n'.format(
					eyeshadow["Name"].replace("'",""), 
					eyeshadow["Brand"].replace("+", ""), 
					eyeshadow["FoundIn"].strip(), 
					eyeshadow["Finish"])
				file.write(eyeshadowline)
			# file.write("};")

	
	def Write_Brands_To_JS(data):
		with open('brands.js', 'w') as file:
			start = "window.brands = [\n"
			file.write(start)
			for brand in data:
				file.write("'{0}'".format(brand["name"].replace(" ", "_").replace("'","").replace("+", "")))
				file.write(",\n")

			file.write("];\n")
			file.close()




	def Write_To_XSLX_Title(fieldnames, file_name):
		workbook = Workbook()
		worksheet = workbook.create_sheet('EyeshadowSheet')


		# workbook.remove_sheet(workbook['Sheet'])
		workbook.remove(workbook['Sheet'])
		row = 1
		col = 1
		for name in fieldnames:
			worksheet.cell(row=row, column = col, value = name)
			col += 1
		workbook.save(file_name)

	def Write_To_XSLX_RGB(data, file_name):
		workbook = openpyxl.load_workbook(filename = file_name)
		worksheet = workbook['EyeshadowSheet']
		for datarow in data:
			col = 1
			row = worksheet.max_row + 1
			for datacell in datarow:
				worksheet.cell(row=row, column=col, value=str(datarow[datacell]))
				col += 1

			# MiddleHSV = datarow["MiddleHSV"]
			# rgb = Color_Class.hsv_to_rgb(MiddleHSV[0],MiddleHSV[1], MiddleHSV[2])
			rgb = datarow["MiddleRGB"]
			color = '{0}'.format(Color_Class.rgb_to_hex(rgb))
			a1 = worksheet[row][col-1]
			a1.fill  = PatternFill(start_color=color, fill_type="solid")

			# AvgHSV = datarow["AvgHSV"]
			# rgb = Color_Class.hsv_to_rgb(AvgHSV[0], AvgHSV[1], AvgHSV[2])
			rgb = datarow["AvgRGB"]
			color = '{0}'.format(Color_Class.rgb_to_hex(rgb))
			
			a2 = worksheet[row][col]
			a2.fill = PatternFill(start_color=color, fill_type="solid")
			
			rgb = datarow["ModeRGB"]
			color = '{0}'.format(Color_Class.rgb_to_hex(rgb))
			a3 = worksheet[row][col+1]
			a3.fill = PatternFill(start_color = color, fill_type="solid")

		#workbook.close()
		workbook.save(file_name)