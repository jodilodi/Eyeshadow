#scrap temptalia brands and check if mongodb has it already, if not insert

from html_scraping import Temptalia_Scrapping
from mongodb import Makeup_MongoDB
from os import system, name
from subprocess import call
import os
import io
from PIL import Image
from array import array
import csv
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import PatternFill
import sys
import math
#~/.bashrc or ~/.bash_aliases
#alias python=python3
#run source ~/.bashrc or source ~/.bash_aliases

def Insert_New_Brands():

	print("Get all brands from Temptalia")
	AllBrands = Temptalia_Scrapping.Get_Brands()
	#print(AllBrands)
	#print(Makeup_MongoDB.Contain_Brand(AllBrands[0]))

	print("Check if brand is already in db and insert")
	for brand in AllBrands:
		print(brand.name)
		brand_exist = Makeup_MongoDB.Contain_Brand(brand.name, brand.id)
		if not brand_exist:
			insertid = Makeup_MongoDB.Insert_Brand(brand.name, brand.id)
			print(insertid)
		else:
			print("exists")

def clear():
	_ = call('clear' if os.name == 'posix' else 'cls')

def Insert_New_Eyeshadows():
	branddata = Makeup_MongoDB.Get_Makeup_DB_After("Makeup Atelier")
	for brand in branddata:
		print(brand)
		totalpages = Temptalia_Scrapping.Get_Nav_Pages(brand["temptalia_id"])
		alleyeshadows = []
		for pageindex in range(1, totalpages + 1):
		#for pageindex in range(1,2):
			alleyeshadows = alleyeshadows + Temptalia_Scrapping.Get_Eyeshadow(brand["name"], brand["temptalia_id"], pageindex)

		for eyeshadow in alleyeshadows:
			exist = Makeup_MongoDB.Contain_Eyeshadow(eyeshadow.brand, eyeshadow.name)
			if not exist:
				insertid = Makeup_MongoDB.Insert_Eyeshadow(eyeshadow)
				print(insertid)
			else:
				print("Added")

def rgb_to_hsv(RGB):
	try:
		r, g, b = RGB[0]/255.0, RGB[1]/255.0, RGB[2]/255.0
	except:
		print ("Error Convert Numbers to RGB_TO_HSV")
		print(RGB)
		return -1
	mx = max(r, g, b)
	mn = min(r, g, b)
	df = mx-mn
	if mx == mn:
	    h = 0
	elif mx == r:
	    h = (60 * ((g-b)/df) + 360) % 360
	elif mx == g:
	    h = (60 * ((b-r)/df) + 120) % 360
	elif mx == b:
	    h = (60 * ((r-g)/df) + 240) % 360
	
	if mx == 0:
	    s = 0
	else:
	    s = (df/mx)*100
	v = mx*100
	return int(round(h)), int(round(s)), int(round(v))

def hsv_to_rgb(h, s, v):
	h = h / 360.0
	s = s / 100.0
	v = v / 100.0
	i = math.floor(h*6)
	f = h*6 - i
	p = v * (1-s)
	q = v * (1-f*s)
	t = v * (1-(1-f)*s)

	r, g, b = [
	    (v, t, p),
	    (q, v, p),
	    (p, v, t),
	    (p, q, v),
	    (t, p, v),
	    (v, p, q),
	][int(i%6)]

	return int(r*255), int(g*255), int(b*255)

def rgb_to_hex(rgb):
 	return '%02x%02x%02x' % rgb

def within_hsv_range(middle, topleft, topright, bottomleft, bottomright):
	threshold = 45
	middlerange = low, high = middle- threshold, middle + threshold
	outsiderange = 0
	
	if ((low > topleft or topleft > high) and (low > (topleft -360) or (topleft - 360) > high)):
		outsiderange += 1
	if ((low > topright or topright > high) and (low > (topright -360) or (topright - 360) > high)):
		outsiderange += 1
	if ((low > bottomleft or bottomleft > high) and (low > (bottomleft -360) or (bottomleft - 360) > high)) :
		outsiderange += 1
	if ((low > bottomright or bottomright > high) and (low > (bottomright -360) or (bottomright - 360) > high)) :
		outsiderange += 1
	
	if outsiderange	> 1:
		return False
	else:
		return True

def Calculate_Avg_HSV(image, middle, borderdistance):
	#This will have a problem is the the eyeshadow is a very shimery eyeshadow and not
	#consistent in the pan
	start = x,y = middle[0] - borderdistance, middle[1] - borderdistance
	#totalr = totalg = totalb = 0
	totalH = totalS = totalV = 0
	for i in range(int(start[0]), int(start[0] + borderdistance*2)):
		for j in range(int(start[1]), int(start[1] + borderdistance*2)):
			pixel = i,j 
			RGB = r,g,b =image.getpixel(pixel)
			HSV = rgb_to_hsv(RGB)
			totalH += HSV[0]
			totalS += HSV[1]
			totalV += HSV[2]
	surfacearea = borderdistance * borderdistance * 4
	return int(totalH/surfacearea), int(totalS/surfacearea), int(totalV/surfacearea)

def Calculate_Mode_HSV(image, middle, borderdistance):
	start = x,y = middle[0] - borderdistance, middle[1] - borderdistance
	HSVDic = {}
	for i in range(int(start[0]), int(start[0] + borderdistance*2)):
		for j in range(int(start[1]), int(start[1] + borderdistance*2)):
			pixel = i,j 
			RGB = r,g,b =image.getpixel(pixel)
			HSV = rgb_to_hsv(RGB)
			if HSV in HSVDic:
				HSVDic[HSV] += 1
			else:
				HSVDic[HSV] = 1
	return len(HSVDic)

def Analyze_Image(image, middle, borderdistance):
	start = x,y = middle[0] - borderdistance, middle[1] - borderdistance
	#totalr = totalg = totalb = 0
	totalH = totalS = totalV = 0
	HSVDic = {}
	for i in range(int(start[0]), int(start[0] + borderdistance*2)):
		for j in range(int(start[1]), int(start[1] + borderdistance*2)):
			pixel = i,j 
			RGB = r,g,b =image.getpixel(pixel)
			HSV = rgb_to_hsv(RGB)
			totalH += HSV[0]
			totalS += HSV[1]
			totalV += HSV[2]
			if HSV in HSVDic:
				HSVDic[HSV] += 1
			else:
				HSVDic[HSV] = 1
	surfacearea = borderdistance * borderdistance * 4
	HSVAvg = H,S,V = int(totalH/surfacearea), int(totalS/surfacearea), int(totalV/surfacearea)
	return HSVAvg, len(HSVDic.keys())

def Write_To_CSV(fieldnames, data ):
	with open('results.csv', 'w', newline='') as csvfile:
		writer = csv.DictWriter(csvfile, fieldnames = fieldnames)
		writer.writeheader()
		for datarow in data:
			writer.writerow(datarow)

def Write_To_XSLX_Title(fieldnames):
	workbook = Workbook()
	worksheet = workbook.create_sheet('EyeshadowSheet')
	workbook.remove_sheet(workbook['Sheet'])
	row = 1
	col = 1
	for name in fieldnames:
		worksheet.cell(row=row, column = col, value = name)
		col += 1
	workbook.save('results.xlsx')

def Write_To_XSLX(data):
	workbook = openpyxl.load_workbook(filename = 'results.xlsx')
	worksheet = workbook['EyeshadowSheet']
	for datarow in data:
		col = 1
		row = worksheet.max_row + 1
		for datacell in datarow:
			worksheet.cell(row=row, column=col, value=str(datarow[datacell]))
			col += 1

		MiddleHSV = datarow["MiddleHSV"]
		rgb = hsv_to_rgb(MiddleHSV[0],MiddleHSV[1], MiddleHSV[2])
		color = '{0}'.format(rgb_to_hex(rgb))
		# cell_format = workbook.add_format()
		# cell_format.set_bg_color(color)
		# worksheet.cell(row = row, column= col, value ="Middle HSV", cell_format)
		a1 = worksheet[row][col-1]
		a1.fill  = PatternFill(start_color=color, fill_type="solid")

		# col += 1

		AvgHSV = datarow["AvgHSV"]
		rgb = hsv_to_rgb(AvgHSV[0], AvgHSV[1], AvgHSV[2])
		color = '{0}'.format(rgb_to_hex(rgb))
		# cell_format1 = workbook.add_format()
		# cell_format1.set_bg_color(color)
		# worksheet.cell(row = row, column = col, value="Avg HSV", cell_format1)
		a2 = worksheet[row][col]
		a2.fill = PatternFill(start_color=color, fill_type="solid")
		
	#workbook.close()
	workbook.save('results.xlsx')
	
def Colors_Within_Range(image, middle, borderdistance, middlehsv):
	results = Analyze_Image(image, middle, borderdistance)
	#print("Middle : {0} vs Avg : {1}".format(middlehsv[0], results[0][0]))
	#if within 60 degree range or when avg + 360 within 60 range
	low = middlehsv[0] - 30
	high = middlehsv[0] + 30
	avghsv = results[0][0]
	withinRange = False
	if (low <= avghsv and avghsv <= high) or (low <= (avghsv + 360) and (avghsv+360) <= high):
		withinRange = True
	return withinRange, results

def Calculate_Image_Box(image, middle, middlehsv):
	borderdistance = 150
	middlex, middley = middle[0], middle[1]
	width, height = image.size

	topleftrgb =image.getpixel((max(1, middlex-borderdistance), max(1, middley - borderdistance)))
	toplefthsv = rgb_to_hsv(topleftrgb)

	toprightrgb = image.getpixel((min(width, middlex+borderdistance),max(1, middley-borderdistance)))
	toprighthsv = rgb_to_hsv(toprightrgb)

	bottomleftrgb = image.getpixel((max(1,middlex-borderdistance), min(height, middley+borderdistance)))
	bottomlefthsv = rgb_to_hsv(bottomleftrgb)

	bottomrightrgb = image.getpixel((min(width, middlex + borderdistance), min(height, middley + borderdistance)))
	bottomrighthsv = rgb_to_hsv(bottomrightrgb)
	#do middle and try for 150 up/down and 150 right/left
	#check if the hue is within 60 degrees if not then go down by 50 all direction
	while not within_hsv_range(middlehsv[0],toplefthsv[0],toprighthsv[0],bottomlefthsv[0],bottomrighthsv[0]) \
		and borderdistance > 0:
		borderdistance -= 5
		topleftrgb =image.getpixel((max(1, middlex-borderdistance), max(1, middley - borderdistance)))
		toplefthsv = rgb_to_hsv(topleftrgb)

		toprightrgb = image.getpixel((min(width, middlex+borderdistance),max(1, middley-borderdistance)))
		toprighthsv = rgb_to_hsv(toprightrgb)

		bottomleftrgb = image.getpixel((max(1,middlex-borderdistance), min(height, middley+borderdistance)))
		bottomlefthsv = rgb_to_hsv(bottomleftrgb)

		bottomrightrgb = image.getpixel((min(width, middlex + borderdistance), min(height, middley + borderdistance)))
		bottomrighthsv = rgb_to_hsv(bottomrightrgb)

	return borderdistance

def Calculate_HSV():
	Brands = Makeup_MongoDB.Get_All_Brands()
	fieldnames = ["Brand", "FoundIn", "Name", "MiddleHSV", "AvgHSV", "UniquePixels", "WithinRange", "MiddleRGB", "AvgRGB"]
	Write_To_XSLX_Title(fieldnames)
	for brand in Brands:
		print(brand)
		Eyeshadows = Makeup_MongoDB.Get_All_Eyeshadows_By_Brand(brand["name"])
		#check four corners and middle
		count = 0
		errorcolors = []
		error = 0
		colorRows = []
		for eyeshadow in Eyeshadows:
			print(eyeshadow["name"])
			try: 
				image = Image.open(io.BytesIO(eyeshadow["byte"]))
			except:
				count+= 1
				errorcolors.append(eyeshadow["name"])
				continue	

			width, height = image.size
			middle = middlex, middley = height/2, width/2
			middlehsv = rgb_to_hsv(image.getpixel(middle))
			if middlehsv == -1:
				error += 1
				continue
			borderdistance = Calculate_Image_Box(image, middle, middlehsv)

			if borderdistance == 0:
				count += 1
				errorcolors.append(eyeshadow["name"])
			else:
				withinRange, results = Colors_Within_Range(image, middle, borderdistance, middlehsv)

				RowDict = {}
				RowDict["Brand"] = eyeshadow["brand"]
				RowDict["FoundIn"] = eyeshadow["foundin"]
				RowDict["Name"] = eyeshadow["name"]
				RowDict["MiddleHSV"] = middlehsv
				RowDict["AvgHSV"] = results[0]
				RowDict["UniquePixels"] = results[1]
				RowDict["WithinRange"] = withinRange
				colorRows.append(RowDict)
		print('writing brand to excel')	
		Write_To_XSLX(colorRows)
	print("Total " + str(count))
	print(errorcolors)
	print("Error " + str(error))
	Write_To_CSV(fieldnames, colorRows)

	#workbook.close()

def Welcome_Screen():
	print('###########################################')
	print('##   Welcome to Temptalia Eyeshadow DB   ##')
	print('###########################################')

def Print_Menu():
	print('Menu')
	print('1. Exit')
	print('2. Delete Brands Collection')
	print('3. Screen Scrap Brands')
	print('4. Delete Eyeshadows Colelction')
	print('5. Screen Scrap Eyeshadows')
	print('6. Print All Eyeshadows')
	print('7. Convert Eyeshadow Byte To Img')
	print('8. Test Pixel Colors')
	print('9. Copy Collection')

if __name__ == "__main__":
	clear()
	Welcome_Screen()
	Print_Menu()
	while True:
		user_input = input('What would you like to do? ')
		if user_input == "Menu" or user_input=="menu":
			Print_Menu()
		elif user_input == "1" or user_input == "Exit" or user_input == "exit" or user_input == "quit":
			print('Goodbye!')
			break
		elif user_input == "2":
			Makeup_MongoDB.Delete_Brands_Collection()
			print('Deleted all brands from MongoDB')
		elif user_input == "3":
			print('Inserting Brands')
			Insert_New_Brands()
		elif user_input == "4":
			Makeup_MongoDB.Delete_Eyeshadow_Collection()
		elif user_input == "5":
			Insert_New_Eyeshadows()
		elif user_input == "6":
			Eyeshadows = Makeup_MongoDB.Get_All_Eyeshadows()
			for es in Eyeshadows:
				print(es)
		elif user_input == "7":
		 	Eyeshadows = Makeup_MongoDB.Get_All_Eyeshadows()
		 	#put everything in brands folder first
		 	if not os.path.isdir("Brands"):
		 		os.mkdir("Brands")

		 	for eyeshadow in Eyeshadows:
		 		brand = eyeshadow["brand"]
		 		filepath = "Brands/{0}".format(brand)
		 		if not os.path.isdir(filepath):
		 			os.mkdir(filepath)
		 		try:
		 			image = Image.open(io.BytesIO(eyeshadow["byte"]))
			 		image.save(filepath + "/" + eyeshadow["name"] + ".jpg")
			 		print("Saved " + eyeshadow["name"] + ".jpg")
			 	except:
			 		print("Could not save " + eyeshadow["name"] + ".jpg")
		elif user_input == "8":
			Calculate_HSV()
		elif user_input == "9":
			from_collection = input('From Collection: ')
			to_collection = input('To Collection: ')
			Makeup_MongoDB.Copy_Collection(from_collection, to_collection)
			print('Complete Copying Collection')
		# elif user_input == "reset":
		# 	os.execl(sys.executable, 'python', "Main.py")
		elif user_input == "reset":
			python = sys.executable
			os.execl(python, python, * sys.argv)